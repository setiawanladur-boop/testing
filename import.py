import os
import time
from openpyxl import load_workbook
import gspread
from google.oauth2.service_account import Credentials
import re

# =========================
# GOOGLE SHEETS CONNECT
# =========================
SERVICE_ACCOUNT_FILE = "service_account.json"
SPREADSHEET_NAME = "FarmingAgus"

scopes = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

try:
    creds = Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=scopes
    )
    client = gspread.authorize(creds)
    spreadsheet = client.open(SPREADSHEET_NAME)
    sheet = spreadsheet.sheet1
    print("✔ KONEKSI GOOGLE SHEETS BERHASIL")
except Exception as e:
    print("❌ GAGAL KONEK GOOGLE SHEETS")
    print("Alasan:", e)
    exit()

# =========================
# KONFIGURASI BARIS
# =========================
ROW_START = 21
MAX_LINK = 5

# =========================
# MAPPING ACCOUNT → KOLOM
# =========================
mapping = {
    "wildan": {
        "INSTAGRAM": ["B", "C", "D"],
        "TIKTOK": ["F", "G", "H"],
        "TWITTER": ["J", "K", "L"],
        "Youtube" : ["N", "O", "P"],
        "SnackVideo" : ["R", "S", "T"],
        "Treads" : ["V", "W", "X"],
        "Facebook" : ["Z", "AA", "AB"],
    },
    "monica": {
        "INSTAGRAM": ["AD", "AE", "AF"],
        "TIKTOK": ["AH", "AI", "AI"],
        "TWITTER": ["AL", "AM", "AN"],
        "Youtube" : ["AP", "AQ", "AR"],
        "SnackVideo" : ["AT", "AU", "AV"],
        "Treads" : ["AX", "AY", "AZ"],
        "Facebook" : ["BB", "BC", "BD"],
    },
    "eko": {
        "INSTAGRAM": ["BF", "BG", "BH"],
        "TIKTOK": ["BJ", "BK", "BL"],
        "TWITTER": ["BN", "BO", "BP"],
        "Youtube" : ["BR", "BS", "BT"],
        "SnackVideo" : ["BV", "BW", "BX"],
        "Treads" : ["BZ", "CA", "CB"],
        "Facebook" : ["CD", "CE", "CF"],
    },
    "zahra": {
        "INSTAGRAM": ["CH", "CI", "CJ"],
        "TIKTOK": ["CL", "CM", "CCN"],
        "TWITTER": ["CP", "CQ", "CR"],
        "Youtube" : ["CT", "CU", "CV"],
        "SnackVideo" : ["CX", "CY", "CZ"],
        "Treads" : ["DB", "DC", "DD"],
        "Facebook" : ["DF", "DG", "DH"],
    },
    "kevin": {
        "INSTAGRAM": ["DJ", "DK", "DL"],
        "TIKTOK": ["DN", "DO", "DP"],
        "TWITTER": ["DR", "DS", "ST"],
        "Youtube" : ["DV", "DW", "DX"],
        "SnackVideo" : ["DZ", "EA", "EB"],
        "Treads" : ["ED", "EE", "EF"],
        "Facebook" : ["EH", "EI", "EJ"],
    },
    "rani": {
        "INSTAGRAM": ["EL", "EM", "EN"],
        "TIKTOK": ["EP", "EQ", "ER"],
        "TWITTER": ["ET", "EU", "EV"],
        "Youtube" : ["EX", "EY", "EZ"],
        "SnackVideo" : ["FB", "FC", "FD"],
        "Treads" : ["FF", "FG", "FH"],
        "Facebook" : ["FJ", "FK", "FL"],
    },
    "wildanf": {
        "INSTAGRAM": ["FN", "FO", "FP"],
        "TIKTOK": ["FR", "FS", "FT"],
        "TWITTER": ["FV", "FW", "FX"],
        "Youtube" : ["FZ", "GA", "GB"],
        "SnackVideo" : ["GD", "GE", "GF"],
        "Treads" : ["GH", "GI", "GJ"],
        "Facebook" : ["GL", "GM", "GN"],
    },
    "arif": {
        "INSTAGRAM": ["GP", "GQ", "GR"],
        "TIKTOK": ["GT", "GU", "GV"],
        "TWITTER": ["GX", "GY", "GZ"],
        "Youtube" : ["HB", "HC", "HD"],
        "SnackVideo" : ["HF", "HG", "HH"],
        "Treads" : ["HJ", "HK", "HL"],
        "Facebook" : ["HN", "HO", "HP"],
    }
}

# =========================
# CODE ORI LU (DIPERTAHANKAN)
# =========================
root_folder = "data"
processed_files = set()   # 🔥 penanda file yang sudah diproses

print("\n🟢 SYSTEM AKTIF — MENUNGGU FILE EXCEL BARU")
print("⛔ STOP : CTRL + C\n")

try:
    while True:
        for folder, _, files in os.walk(root_folder):
            for file in files:
                if file.startswith("~$") or not file.endswith(".xlsx"):
                    continue

                path = os.path.join(folder, file)

                # 🔒 skip jika sudah pernah diproses
                if path in processed_files:
                    continue

                nama = os.path.splitext(file)[0].split("_")[0].lower()

                print("\n======================================")
                print("FILE  :", file)
                print("NAMA  :", nama)
                print("FOLDER:", folder)
                print("======================================")

                try:
                    wb = load_workbook(path, data_only=True)
                    ws = wb.active

                    hasil = {
                        "TIKTOK": [],
                        "INSTAGRAM": [],
                        "TWITTER": [],
                        "Youtube" : [],
                        "SnackVideo" : [],
                        "Treads" : [],
                        "Facebook" : []
                    }

                    for row in ws.iter_rows():
                        for cell in row:
                            link = None
                            if cell.hyperlink:
                                link = cell.hyperlink.target
                            elif isinstance(cell.value, str) and cell.value.startswith("http"):
                                link = cell.value.strip()

                            if not link:
                                continue

                            link_lower = link.lower()
                            if "tiktok.com" in link_lower:
                                hasil["TIKTOK"].append(link)

                            elif "instagram.com" in link_lower:
                                hasil["INSTAGRAM"].append(link)

                            elif "x.com" in link_lower or "twitter.com" in link_lower:
                                hasil["TWITTER"].append(link)

                            elif "youtube.com" in link_lower or "youtu.be" in link_lower:
                                hasil["Youtube"].append(link)

                            elif "s.snackvideo.com" in link_lower:
                                hasil["SnackVideo"].append(link)

                            elif "threads.com" in link_lower:
                                hasil["Treads"].append(link)

                            elif "facebook.com" in link_lower or "fb.com" in link_lower:
                                hasil["Facebook"].append(link)

                    if nama not in mapping:
                        print("⚠ ACCOUNT TIDAK ADA DI MAPPING — SKIP")
                        processed_files.add(path)
                        continue

                    # =========================
                    # TEMPEL KE GOOGLE SHEET
                    # =========================
                    for platform, cols in mapping[nama].items():
                        links = hasil.get(platform, [])[:MAX_LINK]

                        for i, link in enumerate(links):
                            row_target = ROW_START + i
                            col = cols[0]
                            sheet.update_acell(f"{col}{row_target}", link)

                    processed_files.add(path)
                    print("✔ FILE BERHASIL DIPROSES & DIKUNCI")

                except Exception as e:
                    print("❌ GAGAL MEMBACA FILE")
                    print("Alasan:", e)

        time.sleep(3)  # ⏱ scan ulang tiap 3 detik

except KeyboardInterrupt:
    print("\n⛔ SYSTEM DIHENTIKAN MANUAL (CTRL + C)")