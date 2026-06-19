import customtkinter as ctk
import threading, time, subprocess, os, winsound
import pyperclip, uiautomator2 as u2
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
from tkinter import messagebox

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io

import win32gui, win32con
import re

# ================= KONEKSI HP =================
print("🔌 Menghubungkan ke HP...")
d = u2.connect()
print("✔ HP terdeteksi!")

# ================= EXCEL =================
def export_excel(nama,
                 tiktok_list, ig_list, twitter_list,
                 youtube_list, snack_list, threads_list, facebook_list):

    folder = "data"
    os.makedirs(folder, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    file_excel = os.path.join(folder, f"{nama}_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "data_links"
    row = 1

    def add_title(text):
        nonlocal row
        ws.merge_cells(f"A{row}:C{row}")
        ws[f"A{row}"] = text
        ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        row += 1

    def add_links(data):
        nonlocal row
        for link in data:
            ws.merge_cells(f"A{row}:C{row}")
            ws[f"A{row}"] = link
            ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
            row += 1
        row += 1

    add_title("TIKTOK LINKS")
    add_links(tiktok_list)

    add_title("INSTAGRAM LINKS")
    add_links(ig_list)

    add_title("TWITTER LINKS")
    add_links(twitter_list)

    add_title("YOUTUBE LINKS")
    add_links(youtube_list)

    add_title("SNACKVIDEO LINKS")
    add_links(snack_list)

    add_title("THREADS LINKS")
    add_links(threads_list)

    add_title("FACEBOOK LINKS")
    add_links(facebook_list)

    wb.save(file_excel)
    winsound.Beep(1200, 300)
    winsound.Beep(1200, 300)

    log(f"📁 Export sukses: {file_excel}")
    return file_excel

# ================= UI =================
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

app = ctk.CTk()
app.title("🔥 LINK COLLECTOR PRO — ALL PLATFORM")
app.geometry("1300x760")

# ================= DATA =================
tiktok_list, ig_list, twitter_list = [], [], []
youtube_list, snack_list, threads_list, facebook_list = [], [], [], []

running = False
scrcpy_hwnd = None

# ================= UI LAYOUT =================
main = ctk.CTkFrame(app)
main.pack(fill="both", expand=True)

left = ctk.CTkFrame(main, width=340)
left.pack(side="left", fill="y", padx=10, pady=10)

right = ctk.CTkFrame(main)
right.pack(side="right", fill="both", expand=True, padx=10, pady=10)

# ================= INPUT =================
def labeled_entry(parent, text):
    ctk.CTkLabel(parent, text=text).pack(anchor="w", padx=10)
    e = ctk.CTkEntry(parent)
    e.pack(fill="x", padx=10, pady=5)
    return e

entry_nama = labeled_entry(left, "Nama File")
entry_max  = labeled_entry(left, "Max Semua Sosmed")
entry_tanggal = labeled_entry(left, "Tanggal Download (YYYY-MM-DD)")
entry_tanggal.insert(0, datetime.now().strftime("%Y-%m-%d"))

# ================= COUNTER =================
counter_label = ctk.CTkLabel(left,
    text="TT:0 | IG:0 | TW:0 | YT:0 | SN:0 | TH:0 | FB:0")
counter_label.pack(pady=10)

# ================= LOG =================
log_box = ctk.CTkTextbox(left, height=200)
log_box.pack(fill="both", expand=True, padx=10, pady=10)

def log(text):
    log_box.insert("end", text + "\n")
    log_box.see("end")

# ================= SCRCPY =================
scrcpy_frame = ctk.CTkFrame(right)
scrcpy_frame.pack(fill="both", expand=True)

def embed_scrcpy():
    global scrcpy_hwnd

    subprocess.Popen(
        'scrcpy --keyboard=uhid --mouse=uhid --no-audio --max-size 1024 --video-bit-rate 2M --window-title "SCRCPY_EMBED"',
        shell=True
    )

    for _ in range(20):
        time.sleep(0.3)
        hwnd = win32gui.FindWindow(None, "SCRCPY_EMBED")
        if hwnd:
            break
    else:
        log("❌ scrcpy gagal")
        return

    win32gui.SetParent(hwnd, scrcpy_frame.winfo_id())
    scrcpy_hwnd = hwnd

# ================= MONITOR =================
def monitor_clipboard():
    global running
    last_clip = ""

    max_all = int(entry_max.get())

    embed_scrcpy()
    pyperclip.copy("")

    while running:
        clip = pyperclip.paste()

        if clip:
            clip = clip.strip()
        else:
            clip = ""

        # log(f"DEBUG: {repr(clip)}")

        if clip != last_clip and "http" in clip:

            if "tiktok.com" in clip and clip not in tiktok_list and len(tiktok_list) < max_all:
                tiktok_list.append(clip)
                log(f"✓ TikTok {len(tiktok_list)}")

            elif "instagram.com" in clip and clip not in ig_list and len(ig_list) < max_all:
                ig_list.append(clip)
                log(f"✓ IG {len(ig_list)}")

            elif ("twitter.com" in clip or "x.com" in clip) and clip not in twitter_list and len(twitter_list) < max_all:
                twitter_list.append(clip)
                log(f"✓ Twitter {len(twitter_list)}")

            elif ("youtube.com" in clip or "youtu.be" in clip) and clip not in youtube_list and len(youtube_list) < max_all:
                youtube_list.append(clip)
                log(f"✓ YouTube {len(youtube_list)}")

            elif "s.snackvideo.com" in clip.lower() and len(snack_list) < max_all:
                url = extract_url(clip)
                if url and url not in snack_list:
                    snack_list.append(url)
                    log(f"✓ Snack {len(snack_list)}")

            elif "threads.com" in clip and clip not in threads_list and len(threads_list) < max_all:
                threads_list.append(clip)
                log(f"✓ Threads {len(threads_list)}")

            elif ("facebook.com" in clip or "fb.watch" in clip) and clip not in facebook_list and len(facebook_list) < max_all:
                facebook_list.append(clip)
                log(f"✓ Facebook {len(facebook_list)}")

            counter_label.configure(
                text=f"TT:{len(tiktok_list)} | IG:{len(ig_list)} | TW:{len(twitter_list)} | "
                    f"YT:{len(youtube_list)} | SN:{len(snack_list)} | TH:{len(threads_list)} | FB:{len(facebook_list)}"
            )

            last_clip = clip

        time.sleep(0.3)

# ================= BUTTON =================
def start():
    global running

    if not entry_max.get().isdigit():
        messagebox.showerror("ERROR", "Max harus angka!")
        return

    if running:
        return

    # reset semua
    tiktok_list.clear()
    ig_list.clear()
    twitter_list.clear()
    youtube_list.clear()
    snack_list.clear()
    threads_list.clear()
    facebook_list.clear()

    running = True
    log("🚀 START")

    threading.Thread(target=monitor_clipboard, daemon=True).start()

def stop():
    global running
    running = False
    os.system("taskkill /IM scrcpy.exe /F")
    log("⛔ STOP")

def manual_export():
    export_excel(
        entry_nama.get(),
        tiktok_list, ig_list, twitter_list,
        youtube_list, snack_list, threads_list, facebook_list
    )

def download_excel():
    try:
        SERVICE_ACCOUNT_FILE = "service_account.json"
        SPREADSHEET_ID = "1wISU-F5gBxFOQI9JYQSf3uzUgUGnkwUM4MzjJlHcbn8"
        SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

        folder = "hasil Farming"
        os.makedirs(folder, exist_ok=True)

        tanggal = entry_tanggal.get().strip()

        if not tanggal:
            tanggal = datetime.now().strftime("%Y-%m-%d")

        # nama file hanya dari tanggal
        full_nama = f"{tanggal}.xlsx"

        output_path = os.path.join("hasil Farming", full_nama)

        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE,
            scopes=SCOPES
        )

        drive_service = build('drive', 'v3', credentials=creds)

        request = drive_service.files().export_media(
            fileId=SPREADSHEET_ID,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        fh = io.FileIO(output_path, "wb")
        downloader = MediaIoBaseDownload(fh, request)

        done = False
        log("📥 Download mulai...")

        while not done:
            status, done = downloader.next_chunk()
            if status:
                log(f"🔹 Progress: {int(status.progress() * 100)}%")

        log(f"✅ Selesai! File: {output_path}")
        winsound.Beep(1000, 200)

    except Exception as e:
        log(f"❌ Error: {e}")


def extract_url(text):
    match = re.search(r'https?://\S+', text)
    return match.group(0) if match else None


ctk.CTkButton(left, text="START", command=start).pack(fill="x", padx=10, pady=5)
ctk.CTkButton(left, text="STOP", command=stop).pack(fill="x", padx=10, pady=5)
ctk.CTkButton(left, text="EXPORT", command=manual_export).pack(fill="x", padx=10, pady=5)
ctk.CTkButton(
    left,
    text="DOWNLOAD",
    command=lambda: threading.Thread(target=download_excel, daemon=True).start()
).pack(fill="x", padx=10, pady=5)

app.mainloop()